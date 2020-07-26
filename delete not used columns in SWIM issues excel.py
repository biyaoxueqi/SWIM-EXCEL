import openpyxl
# 通过输入excel 地址和名字方式打开工作表格
path = input("please input the excel address:")
filename = input("please input the excel name:")
wb = openpyxl.load_workbook(path+'\\'+filename)
wsname = input("please input the worksheet name:")
ws = wb[wsname]

list1 =["Key"]
rowidx=1
colidx=1
rownum=1
colnum=1
# 找到Key所在列
for i in range(1,5):
    for j in range(1,10):
        if ws.cell(i,j).value in list1:
            colidx = j
            rowidx = i

# 选择要保留的列
list2 = ["Key","ECU","Status","Created","Risk Factor","Responsible Project","Target Serie","Found in Serie"]
while  colnum < ws.max_column+1:
    if ws.cell(rowidx,colnum).value not in list2:
        print(ws.cell(rowidx,colnum).value)
        ws.delete_cols(colnum)


    else:
        colnum +=1
#选择是存在原有表格还是新表格
oldpath = input("if you want to save the excel in old path or not(type Y or N):")
if oldpath == "Y":
    savepath = path+"\\"+filename
else:
    savepath=input("please input the save excel path and name:")
wb.save(savepath)